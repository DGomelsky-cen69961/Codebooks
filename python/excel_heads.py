import openpyxl

workbook = openpyxl.load_workbook('TEST_Colman_Code_Tables_Common.xlsx')
with open('output.txt', 'w') as file:
    for worksheet in workbook.worksheets:
        worksheet_name = worksheet.title
        max_column = worksheet.max_column
        file.write(str(worksheet_name))

        for column in range(1, max_column + 1):
            cell_value = worksheet.cell(row=1, column=column).value
            file.write(str(cell_value + '; '))

        file.write(str('\n'))